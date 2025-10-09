import streamlit as st
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

# ==============================================================================
#  INÍCIO: Lógica de conversão e formatação (adaptada do seu script original)
# ==============================================================================

# Função para formatar o Workbook do Excel em memória
def formatar_excel_workbook(wb):
    """
    Aplica formatação a um objeto workbook do openpyxl.
    - Ajusta largura das colunas
    - Define zoom
    - Adiciona filtros
    - Valida preenchimento de horas, colorindo linhas incompletas
    """
    ws = wb.active

    # Ajusta largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Ajusta um pouco mais para dar um respiro
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Define zoom 75%
    ws.sheet_view.zoomScale = 75

    # Adiciona filtro automático na primeira linha
    ws.auto_filter.ref = ws.dimensions

    # Colunas alvo para validar preenchimento
    colunas_alvo = ["1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S."]

    # Descobre os índices das colunas alvo
    try:
        header = [cell.value for cell in ws[1]]
        idx_dia = header.index("Dia") + 1
        idx_1aE = header.index("1a E.") + 1
        idx_alvo = [header.index(c) + 1 for c in colunas_alvo if c in header]
    except ValueError as e:
        # Se uma coluna essencial não for encontrada, pula a formatação condicional
        st.warning(f"Atenção: A coluna '{e.args[0].split(' is not in list')[0]}' não foi encontrada. A validação de preenchimento será ignorada.")
        return wb

    # Estilo de preenchimento vermelho claro
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Percorre linhas (a partir da 2ª) para aplicar formatação condicional
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        dia = row[idx_dia - 1].value
        col_1aE = row[idx_1aE - 1].value

        # Ignora sábados, domingos e ausentes
        if dia and isinstance(dia, str) and dia.strip().lower() in ["sábado", "sabado", "domingo"]:
            continue
        if col_1aE == "** Ausente **":
            continue

        # Conta células preenchidas nas colunas de marcação
        preenchidos = sum(1 for i in idx_alvo if row[i - 1].value not in [None, ""])
        if preenchidos > 0 and preenchidos < 4: # Considera incompleto se tiver alguma marcação mas não todas as 4
            for cell in row:
                cell.fill = red_fill
    
    return wb


# Função principal para extrair dados do PDF e retornar um arquivo Excel em bytes
def converter_pdf_para_bytes_excel(pdf_file):
    """
    Lê um arquivo PDF, extrai os dados do espelho de ponto e retorna
    um arquivo Excel formatado como um objeto de bytes.
    """
    dados = []
    colunas = None
    nome_atual, matricula_atual, funcao_atual, turno_atual = None, None, None, None

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()

            # Tenta extrair informações do cabeçalho da página
            for linha in texto.split("\n"):
                if "Nome:" in linha:
                    try:
                        matricula_atual = linha.split("Matrícula:")[1].split("Nome:")[0].strip()
                        nome_atual = linha.split("Nome:")[1].split("Chapa:")[0].strip()
                    except IndexError:
                        continue # Ignora linhas mal formatadas
                if "Função:" in linha:
                    try:
                        funcao_atual = linha.split("Função:")[1].strip()
                    except IndexError:
                        continue
                if "Turno:" in linha:
                    try:
                        turno_atual = linha.split("Turno:")[1].strip()
                    except IndexError:
                        continue

            # Extrai tabelas da página
            tabelas = page.extract_tables()
            for tabela in tabelas:
                # Verifica se é a tabela de ponto (procurando por 'Data' no cabeçalho)
                if tabela and "Data" in tabela[0]:
                    if colunas is None:
                        colunas = ["Nome", "Matrícula", "Função", "Turno"] + tabela[0]
                    for linha in tabela[1:]:
                        # Ignora linhas completamente vazias
                        if any(cel for cel in linha):
                            dados.append([nome_atual, matricula_atual, funcao_atual, turno_atual] + linha)

    if not dados:
        st.error("Nenhuma tabela de ponto válida foi encontrada no PDF.")
        return None

    # Monta o DataFrame com os dados extraídos
    df = pd.DataFrame(dados, columns=colunas)

    # --- Salva o DataFrame em um buffer de memória para não criar arquivo em disco ---
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Espelho de Ponto')
        
        # Carrega o workbook a partir do writer para aplicar formatação
        workbook = writer.book
        formatar_excel_workbook(workbook)

    # Pega o conteúdo do buffer (o arquivo Excel em si)
    excel_bytes = output_buffer.getvalue()
    return excel_bytes


# ==============================================================================
#  FIM: Lógica de conversão
# ==============================================================================


# --- Interface Gráfica com Streamlit ---

st.set_page_config(page_title="Conversor de Espelho de Ponto", layout="centered")

st.title("Conversor de Espelho de Ponto 📄➡️📊")
st.write("Anexe o arquivo PDF do espelho de ponto para convertê-lo para um arquivo Excel formatado.")

# 1. Campo para anexar arquivos PDF
uploaded_file = st.file_uploader(
    "Escolha o arquivo PDF",
    type="pdf",
    label_visibility="collapsed"
)

# Verifica se um arquivo foi carregado
if uploaded_file is not None:
    
    # Gera um nome de arquivo padrão para o download
    default_file_name = uploaded_file.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')
    
    # 2. Botão "Converter"
    if st.button("Converter para Excel", type="primary"):
        
        # Mostra uma mensagem de "processando"
        with st.spinner("Aguarde... Convertendo e formatando o arquivo..."):
            try:
                # Chama a função principal que processa o PDF
                excel_data = converter_pdf_para_bytes_excel(uploaded_file)

                if excel_data:
                    st.success("Arquivo convertido com sucesso! 🎉")
                    
                    # 3. Botão de Download (que substitui a caixa de "Salvar como")
                    st.download_button(
                        label="Clique aqui para baixar o Excel",
                        data=excel_data,
                        file_name=default_file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Ocorreu um erro inesperado durante a conversão: {e}")
                st.error("Por favor, verifique se o arquivo PDF tem o formato esperado.")